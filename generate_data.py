"""
Generate fresh sample Excel data matching the new dashboard design.
Run once: python generate_data.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
os.makedirs(DATA_DIR, exist_ok=True)

H_FILL  = PatternFill("solid", start_color="1B4332")
H_FONT  = Font(bold=True, color="FFFFFF", name="Calibri")
D_FONT  = Font(name="Calibri", size=11)
C_ALIGN = Alignment(horizontal="center", vertical="center")

def style_ws(ws, col_widths):
    for cell in ws[1]:
        cell.font = H_FONT
        cell.fill = H_FILL
        cell.alignment = C_ALIGN
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1: Fields Master
# ─────────────────────────────────────────────────────────────────────────────
wb = Workbook()
ws1 = wb.active
ws1.title = "Fields"
ws1.append(["field_id","field_name","block","crop","season","is_active","target_qty_per_week"])
style_ws(ws1, [10,12,14,14,12,10,22])
fields = [
    ["F001","Field A","North Block","Tomato","March 2026",True,5000],
    ["F002","Field B","South Block","Pepper","March 2026",True,3200],
    ["F003","Field C","East Block","Cucumber","March 2026",True,4500],
    ["F004","Field D","West Block","Lettuce","March 2026",True,6000],
    ["F005","GH-1","Greenhouse 1","Mixed Herbs","March 2026",True,8000],
]
for r in fields:
    ws1.append(r)

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2: Weekly Planting Data (per field per week)
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Weekly_Planting")
ws2.append(["week_number","week_label","season","field_id","field_name","block","crop",
            "planned_date","actual_date","labor_cost","qty_planted","target_qty","notes"])
style_ws(ws2, [13,12,14,10,12,14,14,14,14,13,13,12,20])

weekly_data = [
    # Wk1
    [1,"Week 1","March 2026","F001","Field A","North Block","Tomato","2026-03-03","2026-03-03",3500,4800,5000,""],
    [1,"Week 1","March 2026","F002","Field B","South Block","Pepper","2026-03-03","2026-03-03",2800,2300,3200,"Late start"],
    [1,"Week 1","March 2026","F003","Field C","East Block","Cucumber","2026-03-03","2026-03-03",5000,4200,4500,""],
    [1,"Week 1","March 2026","F004","Field D","West Block","Lettuce","2026-03-03","2026-03-10",4500,3400,6000,"1 week behind"],
    [1,"Week 1","March 2026","F005","GH-1","Greenhouse 1","Mixed Herbs","2026-03-03","2026-03-10",5500,7000,8000,""],
    # Wk2
    [2,"Week 2","March 2026","F001","Field A","North Block","Tomato","2026-03-10","2026-03-10",3500,4900,5000,""],
    [2,"Week 2","March 2026","F002","Field B","South Block","Pepper","2026-03-10","2026-03-10",2800,2450,3200,""],
    [2,"Week 2","March 2026","F003","Field C","East Block","Cucumber","2026-03-10","2026-03-10",5000,4300,4500,""],
    [2,"Week 2","March 2026","F004","Field D","West Block","Lettuce","2026-03-10","2026-03-17",4500,3500,6000,"Still behind"],
    [2,"Week 2","March 2026","F005","GH-1","Greenhouse 1","Mixed Herbs","2026-03-10","2026-03-10",5500,7100,8000,""],
    # Wk3 (current)
    [3,"Week 3","March 2026","F001","Field A","North Block","Tomato","2026-03-17","2026-03-17",3500,4850,5000,""],
    [3,"Week 3","March 2026","F002","Field B","South Block","Pepper","2026-03-17","2026-03-24",2800,2400,3200,"1 wk behind"],
    [3,"Week 3","March 2026","F003","Field C","East Block","Cucumber","2026-03-17","2026-03-17",5000,4350,4500,""],
    [3,"Week 3","March 2026","F004","Field D","West Block","Lettuce","2026-03-17","2026-04-07",4500,3600,6000,"3 wks behind"],
    [3,"Week 3","March 2026","F005","GH-1","Greenhouse 1","Mixed Herbs","2026-03-17","2026-03-24",5500,7200,8000,"1 wk behind"],
]
for r in weekly_data:
    ws2.append(r)

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3: KPI Thresholds (configurable)
# ─────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Thresholds")
ws3.append(["kpi_name","green_max","orange_max","unit","description"])
style_ws(ws3, [22,12,12,10,40])
ws3.append(["cost_per_seedling",2.50,3.50,"R","Cost per seedling in Rands"])
ws3.append(["qty_achievement_pct",95,80,"%","Actual qty as % of target"])
ws3.append(["weeks_behind",0,1,"weeks","Weeks behind planting schedule"])

wb.save(os.path.join(DATA_DIR, "nuharvest_planting.xlsx"))
print("✅ nuharvest_planting.xlsx created in data/")

# ─────────────────────────────────────────────────────────────────────────────
# Sample upload file (for the uploader)
# ─────────────────────────────────────────────────────────────────────────────
wb2 = Workbook()
ws_up = wb2.active
ws_up.title = "Weekly_Planting"
ws_up.append(["week_number","week_label","season","field_id","field_name","block","crop",
              "planned_date","actual_date","labor_cost","qty_planted","target_qty","notes"])
style_ws(ws_up, [13,12,14,10,12,14,14,14,14,13,13,12,20])
upload_rows = [
    [4,"Week 4","March 2026","F001","Field A","North Block","Tomato","2026-03-24","2026-03-24",3500,4900,5000,""],
    [4,"Week 4","March 2026","F002","Field B","South Block","Pepper","2026-03-24","2026-03-24",2800,2600,3200,"Caught up"],
    [4,"Week 4","March 2026","F003","Field C","East Block","Cucumber","2026-03-24","2026-03-24",5000,4400,4500,""],
    [4,"Week 4","March 2026","F004","Field D","West Block","Lettuce","2026-03-24","2026-03-31",4500,4000,6000,"2 wks behind"],
    [4,"Week 4","March 2026","F005","GH-1","Greenhouse 1","Mixed Herbs","2026-03-24","2026-03-24",5500,7800,8000,""],
]
for r in upload_rows:
    ws_up.append(r)
wb2.save(os.path.join(DATA_DIR, "sample_week4_upload.xlsx"))
print("✅ sample_week4_upload.xlsx created in data/")
