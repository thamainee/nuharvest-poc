"""
kpi_reader.py — reads from DB, calculates KPIs, generates notifications
"""
from django.conf import settings
from .models import WeeklyPlanting, Department, Notification, User


STATUS_COLOUR = {"GREEN": "#2D9B56", "ORANGE": "#F4A261", "RED": "#E63946"}


def get_dashboard_data(department=None, season=None):
    """
    Returns all KPI data for the dashboard.
    department: Department object or None (returns all)
    season: string like "March 2026" or None (latest)
    """
    qs = WeeklyPlanting.objects.all()
    if department:
        qs = qs.filter(department=department)

    # Get available seasons
    seasons = list(qs.values_list("season", flat=True).distinct().order_by("season"))
    if not season and seasons:
        season = seasons[-1]
    if season:
        qs = qs.filter(season=season)

    # Get latest week
    weeks = sorted(qs.values_list("week_number", flat=True).distinct())
    current_week = weeks[-1] if weeks else 0
    total_weeks  = 8  # season length

    current_qs = qs.filter(week_number=current_week)
    records     = list(current_qs.order_by("field_name"))

    # ── Per-field KPIs ─────────────────────────────────────────────────────
    field_rows = []
    for r in records:
        field_rows.append({
            "field_name":        r.field_name,
            "block":             r.block,
            "crop":              r.crop,
            "cost_per_seedling": r.cost_per_seedling,
            "cost_status":       r.cost_status,
            "cost_colour":       STATUS_COLOUR[r.cost_status],
            "schedule_label":    r.schedule_label,
            "schedule_status":   r.schedule_status,
            "schedule_colour":   STATUS_COLOUR[r.schedule_status],
            "achievement_pct":   r.achievement_pct,
            "qty_status":        r.qty_status,
            "qty_colour":        STATUS_COLOUR[r.qty_status],
            "target_qty":        r.target_qty,
            "qty_planted":       r.qty_planted,
            "labor_cost":        r.labor_cost,
            "weeks_behind":      r.weeks_behind,
            "planned_date":      str(r.planned_date) if r.planned_date else "",
            "actual_date":       str(r.actual_date)  if r.actual_date  else "",
            "notes":             r.notes,
        })

    # ── Summary KPIs ───────────────────────────────────────────────────────
    active_fields      = len(records)
    fields_on_schedule = sum(1 for r in records if r.schedule_status == "GREEN")
    total_qty_planted  = sum(r.qty_planted  for r in records)
    total_qty_target   = sum(r.target_qty   for r in records)
    total_labor        = sum(r.labor_cost   for r in records)
    overall_pct        = round(total_qty_planted / total_qty_target * 100, 1) if total_qty_target else 0
    avg_cost           = round(total_labor / total_qty_planted, 2) if total_qty_planted else 0

    # ── Historical trend (all weeks, per field) ────────────────────────────
    all_records = list(qs.order_by("week_number", "field_name"))
    field_names = sorted(set(r.field_name for r in all_records))
    week_labels = [f"Wk {w}" for w in sorted(weeks)]

    # cost trend per field per week
    cost_trend = {}
    qty_trend  = {}
    for fn in field_names:
        cost_trend[fn] = []
        qty_trend[fn]  = []
        for wk in sorted(weeks):
            match = next((r for r in all_records if r.week_number == wk and r.field_name == fn), None)
            cost_trend[fn].append(round(match.cost_per_seedling, 2) if match else None)
            qty_trend[fn].append(match.qty_planted if match else None)

    # Actual vs target for current week (bar chart)
    bar_labels  = [r["field_name"] for r in field_rows]
    bar_targets = [r["target_qty"]  for r in field_rows]
    bar_actuals = [r["qty_planted"] for r in field_rows]

    # ── Alerts / notifications ─────────────────────────────────────────────
    alerts = []
    for r in field_rows:
        if r["cost_status"] in ("ORANGE", "RED"):
            alerts.append({
                "severity":  r["cost_status"],
                "colour":    r["cost_colour"],
                "title":     f"{r['field_name']} — Cost Alert",
                "message":   f"Cost per seedling is R{r['cost_per_seedling']} ({r['cost_status']})",
                "field":     r["field_name"],
                "kpi":       "Cost/Seedling",
                "value":     f"R{r['cost_per_seedling']}",
            })
        if r["schedule_status"] in ("ORANGE", "RED"):
            alerts.append({
                "severity":  r["schedule_status"],
                "colour":    r["schedule_colour"],
                "title":     f"{r['field_name']} — Schedule Alert",
                "message":   f"Planting is {r['schedule_label']} ({r['schedule_status']})",
                "field":     r["field_name"],
                "kpi":       "Schedule",
                "value":     r["schedule_label"],
            })
        if r["qty_status"] in ("ORANGE", "RED"):
            alerts.append({
                "severity":  r["qty_status"],
                "colour":    r["qty_colour"],
                "title":     f"{r['field_name']} — Qty Alert",
                "message":   f"Achievement is {r['achievement_pct']}% of target ({r['qty_status']})",
                "field":     r["field_name"],
                "kpi":       "Qty Achievement",
                "value":     f"{r['achievement_pct']}%",
            })

    alerts.sort(key=lambda a: 0 if a["severity"] == "RED" else 1)
    has_red    = any(a["severity"] == "RED"    for a in alerts)
    has_orange = any(a["severity"] == "ORANGE" for a in alerts)

    return {
        "season":          season,
        "seasons":         seasons,
        "current_week":    current_week,
        "total_weeks":     total_weeks,
        "week_label":      f"Week {current_week} of {total_weeks}",
        "active_fields":   active_fields,
        "fields_on_schedule": fields_on_schedule,
        "total_fields":    active_fields,
        "total_qty_planted": total_qty_planted,
        "total_qty_target":  total_qty_target,
        "overall_pct":     overall_pct,
        "avg_cost":        avg_cost,
        "total_labor":     total_labor,
        "field_rows":      field_rows,
        "field_names":     field_names,
        "week_labels":     week_labels,
        "cost_trend":      cost_trend,
        "qty_trend":       qty_trend,
        "bar_labels":      bar_labels,
        "bar_targets":     bar_targets,
        "bar_actuals":     bar_actuals,
        "alerts":          alerts,
        "has_red":         has_red,
        "has_orange":      has_orange,
        "alert_count":     len(alerts),
    }


def load_excel_to_db(department):
    """Load the Excel file into the DB for the given department."""
    import openpyxl
    from datetime import datetime

    path = settings.PLANTING_EXCEL_PATH
    wb   = openpyxl.load_workbook(path, data_only=True)

    def parse_date(v):
        if v is None: return None
        if hasattr(v, "date"): return v.date() if callable(v.date) else v
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(str(v).strip(), fmt).date()
            except: pass
        return None

    ws = wb["Weekly_Planting"]
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

    created = updated = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row): continue
        d = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        _, was_created = WeeklyPlanting.objects.update_or_create(
            week_number  = int(d.get("week_number", 0)),
            season       = str(d.get("season", "")).strip(),
            field_id_ref = str(d.get("field_id", "")).strip(),
            defaults={
                "department":   department,
                "week_label":   str(d.get("week_label", "")).strip(),
                "field_name":   str(d.get("field_name", "")).strip(),
                "block":        str(d.get("block", "")).strip(),
                "crop":         str(d.get("crop", "")).strip(),
                "planned_date": parse_date(d.get("planned_date")),
                "actual_date":  parse_date(d.get("actual_date")),
                "labor_cost":   float(d.get("labor_cost") or 0),
                "qty_planted":  int(float(str(d.get("qty_planted") or 0))),
                "target_qty":   int(float(str(d.get("target_qty") or 0))),
                "notes":        str(d.get("notes") or "").strip(),
            }
        )
        if was_created: created += 1
        else: updated += 1

    wb.close()
    return created, updated
